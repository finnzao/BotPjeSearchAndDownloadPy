from selenium import webdriver
from selenium.webdriver.support.expected_conditions import StaleElementReferenceException
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support.ui import Select
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
from openpyxl import Workbook
from openpyxl.styles import Font
import time
from dotenv import load_dotenv
import os
import math

def initialize_driver():
    driver = webdriver.Chrome()
    wait = WebDriverWait(driver, 20)
    return driver, wait

def login(driver, wait, user, password):
    login_url = 'https://pje.tjba.jus.br/pje/login.seam'
    driver.get(login_url)
    wait.until(EC.frame_to_be_available_and_switch_to_it((By.ID, 'ssoFrame')))
    username_input = wait.until(EC.presence_of_element_located((By.ID, 'username')))
    password_input = wait.until(EC.presence_of_element_located((By.ID, 'password')))
    login_button = wait.until(EC.presence_of_element_located((By.ID, 'kc-login')))
    username_input.send_keys(user)
    password_input.send_keys(password)
    login_button.click()
    driver.switch_to.default_content()

def skip_token(driver, wait):
    proceed_button = wait.until(
        EC.element_to_be_clickable((By.XPATH, "//a[contains(text(),'Prosseguir sem o Token')]"))
    )
    proceed_button.click()

def select_profile(driver, wait, profile):
    dropdown = wait.until(EC.presence_of_element_located((By.CLASS_NAME, 'dropdown-toggle')))
    dropdown.click()

    button_xpath = f"//a[contains(text(), '{profile}')]"
    
    desired_button = wait.until(EC.element_to_be_clickable((By.XPATH, button_xpath)))

    try:
        wait.until(EC.invisibility_of_element_located((By.CLASS_NAME, 'overlay')))
    except:
        pass

    driver.execute_script("arguments[0].scrollIntoView(true);", desired_button)
    driver.execute_script("arguments[0].click();", desired_button)

def search_process(driver, wait, classeJudicial:str='', nomeParte:str='',numOrgaoJustica:int='0216',numeroOAB:int='',estadoOAB:int=''):
    wait.until(EC.frame_to_be_available_and_switch_to_it((By.ID, 'ngFrame')))
    icon_search_button = wait.until(
        EC.element_to_be_clickable((By.CSS_SELECTOR, 'li#liConsultaProcessual i.fas'))
    ) 
    icon_search_button.click()
    wait.until(EC.frame_to_be_available_and_switch_to_it((By.ID, 'frameConsultaProcessual')))
    
    ElementoNumOrgaoJutica = wait.until(EC.presence_of_element_located((By.ID, 'fPP:numeroProcesso:NumeroOrgaoJustica')))
    ElementoNumOrgaoJutica.send_keys(numOrgaoJustica)
   
    # OAB
    if estadoOAB:
        ElementoNumeroOAB = wait.until(EC.presence_of_element_located((By.ID, 'fPP:decorationDados:numeroOAB')))
        ElementoNumeroOAB.send_keys(numeroOAB)
        ElementoEstadosOAB = wait.until(EC.presence_of_element_located((By.ID, 'fPP:decorationDados:ufOABCombo')))
        listaEstadosOAB = Select(ElementoEstadosOAB)
        listaEstadosOAB.select_by_value(estadoOAB) # 4 .:. BA
    
    consulta_classe = wait.until(EC.presence_of_element_located((By.ID, 'fPP:j_id245:classeJudicial')))
    consulta_classe.send_keys(classeJudicial)
    
    ElementonomeDaParte = wait.until(EC.presence_of_element_located((By.ID, 'fPP:j_id150:nomeParte')))
    ElementonomeDaParte.send_keys(nomeParte)
    
    btnProcurarProcesso = wait.until(EC.presence_of_element_located((By.ID, 'fPP:searchProcessos')))
    btnProcurarProcesso.click()

def collect_process_numbers(driver, wait):
    WebDriverWait(driver, 50).until(
        EC.presence_of_element_located((By.CSS_SELECTOR, "div.pull-right span.text-muted"))
    )
    
    total_resultados = None
    
    while total_resultados is None:
        try:
            total_resultados_texto = driver.find_element(By.CSS_SELECTOR, "div.pull-right span.text-muted").text
            total_resultados = int(total_resultados_texto.split()[0])
        except (ValueError, IndexError):
            continue
    
    itens_por_pagina = 20
    total_paginas = math.ceil(total_resultados / itens_por_pagina)
    
    numProcessos = []

    for pagina in range(total_paginas):
        WebDriverWait(driver, 50).until(
            EC.visibility_of_all_elements_located((By.CSS_SELECTOR, "a.btn-link.btn-condensed"))
        )

        links_dos_processos = driver.find_elements(By.CSS_SELECTOR, "a.btn-link.btn-condensed")

        for link in links_dos_processos:
            numero_do_processo = link.get_attribute('title')
            numProcessos.append(numero_do_processo)

        if pagina == total_paginas - 1:
            print("Fim da paginação")
            break

        try:
            next_page_button = wait.until(
                EC.element_to_be_clickable((By.XPATH, "//td[contains(@onclick, 'next')]"))
            )
            next_page_button.click()
        except Exception as e:
            print("Erro ao tentar clicar no botão de próxima página:", e)
            break

    numUnico = set(numProcessos)
    numUnicosLista = list(numUnico)
    return numUnicosLista

def save_to_excel(process_numbers, filename="ResultadoProcessosPesquisa"):
    dir_path = "./docs"
    file_path = f"{dir_path}/{filename}.xlsx"
    
    if not os.path.exists(dir_path):
        os.makedirs(dir_path)
    
    wb = Workbook()
    ws = wb.active
    ws.title = filename
    bold_font = Font(bold=True, size=16)
    ws["A1"] = "Processos"
    ws["A1"].font = bold_font
    
    for row, processo in enumerate(process_numbers, start=2):
        ws[f"A{row}"] = processo

    wb.save(filename=file_path)
    print(f"Arquivo '{file_path}' criado com sucesso.")

def nav_tag(driver, wait):
    tag_button = wait.until(
        EC.element_to_be_clickable((By.CSS_SELECTOR, 'li#liEtiquetas i.fas'))
    ) 
    tag_button.click()

def search_on_tag(driver, wait, search):
    wait.until(EC.frame_to_be_available_and_switch_to_it((By.ID, 'ngFrame')))
    nav_tag(driver, wait)
    input_tag(driver, wait, search)
    click_element_on_result(driver, wait)
    click_element_by_selector(css_selector='a.selecionarProcesso',driver=driver,max_retries=2,wait=wait)
def input_tag(driver, wait, search_text):
    search_input = wait.until(EC.element_to_be_clickable((By.ID, "itPesquisarEtiquetas")))
    search_input.clear()
    search_input.send_keys(search_text)
    search_button = driver.find_element(By.CSS_SELECTOR, "button.btn.btn-default.btn-pesquisa[title='Pesquisar']")
    search_button.click()

def click_element_on_result(driver, wait, max_retries=3):
    #driver.switch_to.default_content()
    #wait.until(EC.frame_to_be_available_and_switch_to_it((By.ID, 'ngFrame')))
    retries = 0
    while retries < max_retries:
        try:
            # Localize o elemento usando o seletor CSS
            target_element = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, 'div.ui-datalist-content .ui-datalist-data .nivel-1 .nome-etiqueta.selecionada')))

            # Rola o elemento para a visualização
            driver.execute_script("arguments[0].scrollIntoView(true);", target_element)

            # Espera até que o elemento esteja clicável
            target_element = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, 'div.ui-datalist-content .ui-datalist-data .nivel-1 .nome-etiqueta.selecionada')))

            # Clica no elemento
            target_element.click()
            break
        except TimeoutException:
            retries += 1
            print(f"Tentativa {retries} falhou. Tentando novamente...")
            if retries >= max_retries:
                raise TimeoutException(f"Não foi possível localizar ou clicar no elemento após {max_retries} tentativas")
        except StaleElementReferenceException:
            retries += 1
            print(f"StaleElementReferenceException capturada. Tentando novamente... ({retries}/{max_retries})")
        except Exception as e:
            print(f"Erro: {e}")
            break
        finally:
            driver.switch_to.default_content()
    if retries == max_retries:
        print(f"Falha ao clicar no elemento após {max_retries} tentativas")


def click_element_by_selector(driver, wait, css_selector, max_retries=3):
    """
    Clica em um elemento localizado pelo seletor CSS.
    
    :param driver: WebDriver instance
    :param wait: WebDriverWait instance
    :param css_selector: Seletor CSS do elemento a ser clicado
    :param max_retries: Número máximo de tentativas em caso de falha
    """
    retries = 0
    while retries < max_retries:
        try:
            # Localiza o elemento usando o seletor CSS
            element = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, css_selector)))
            
            # Rola o elemento para a visualização
            driver.execute_script("arguments[0].scrollIntoView(true);", element)
            
            # Espera até que o elemento esteja clicável
            element = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, css_selector)))
            
            # Clica no elemento
            element.click()
            print("Elemento clicado com sucesso.")
            break  # Sai do loop se o clique for bem-sucedido
            
        except Exception as e:
            retries += 1
            print(f"Tentativa {retries} falhou. Erro: {e}")
            if retries >= max_retries:
                print(f"Não foi possível clicar no elemento após {max_retries} tentativas")
                raise e  # Relança a exceção após o número máximo de tentativas

def main():
    load_dotenv()
    driver, wait = initialize_driver()
    user, password = os.getenv("USER"), os.getenv("PASSWORD")
    profile = os.getenv("PROFILE")
    optionSearch= {'classeJudicial':"Curatela", 'nomeParte':''}
    login(driver, wait, user, password)
    select_profile(driver, wait, "V DOS FEITOS DE REL DE CONS CIV E COMERCIAIS DE RIO REAL / Direção de Secretaria / Diretor de Secretaria")
    search_on_tag(driver, wait, "Cobrar Custas")
    driver.quit()

if __name__ == "__main__":
    main()
